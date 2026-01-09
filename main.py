import os
import re
import json
import time
import hashlib
import threading
import difflib
from datetime import datetime
from io import BytesIO

import telebot
import requests
from openpyxl import load_workbook

# === КОНФИГУРАЦИЯ ===
BOT_TOKEN = "8573758555:AAF5J7AuctgI_7UBOAggtl2jSVFUKYdeG_A"
URL = "https://vsuet.ru/images/student/schedule/uits.xlsx"
CHANNEL_ID = "-1003374839005"  # <--- ЗАМЕНИ НА СВОЙ ID КАНАЛА
POLL_SECONDS = 600            # Интервал проверки (10 мин)
# ====================

USERS_FILE = "users.json"
STATE_FILE = "state.json"
GROUP_PATTERN = re.compile(r"\bуб[-\s]?21\b", re.IGNORECASE)

is_paused = False
pause_event = threading.Event()
pause_event.set()

bot = telebot.TeleBot(BOT_TOKEN)

# --- БАЗА ЮЗЕРОВ (для спама в личку) ---
def load_users():
    if not os.path.exists(USERS_FILE): return []
    try:
        with open(USERS_FILE, "r") as f: return json.load(f)
    except: return []

def save_user(chat_id):
    users = load_users()
    if chat_id not in users:
        users.append(chat_id)
        with open(USERS_FILE, "w") as f: json.dump(users, f)

# --- КОМАНДЫ ---
@bot.message_handler(commands=['start'])
def handle_start(message):
    save_user(message.chat.id)
    bot.send_message(message.chat.id, "✅ Подписка активна. Буду кидать файл каждые 10 минут.")

@bot.message_handler(commands=['go'])
def handle_go(message):
    global is_paused
    if is_paused:
        is_paused = False
        pause_event.set()
        bot.reply_to(message, "▶️ Пауза снята. Мониторинг продолжается.")
        try:
            bot.send_message(CHANNEL_ID, "ℹ️ <i>Мониторинг возобновлен.</i>", parse_mode="HTML")
        except: pass
    else:
        bot.reply_to(message, "Я и не спал.")

# --- УТИЛИТЫ ---
def load_state():
    if not os.path.exists(STATE_FILE): return {"hash": None, "text": None}
    try:
        with open(STATE_FILE, "r") as f: return json.load(f)
    except: return {"hash": None, "text": None}

def save_state(h, text):
    with open(STATE_FILE, "w") as f:
        json.dump({"hash": h, "text": text}, f, ensure_ascii=False, indent=2)

def download_xlsx(url):
    try:
        r = requests.get(url, timeout=60)
        return r.content
    except: return None

def extract_text(content):
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        if len(wb.worksheets) < 4: return ""
        lines = []
        for row in wb.worksheets[3].iter_rows(values_only=True):
            t = " | ".join([str(v).strip() for v in row if v])
            if t and GROUP_PATTERN.search(t): lines.append(t)
        return "\n".join(lines)
    except: return ""

def make_diff(old, new):
    d = list(difflib.unified_diff((old or "").splitlines(), (new or "").splitlines(), lineterm=""))
    return "\n".join(d[:15]) if d else ""

# --- ЯДРО ---
def watcher_loop():
    global is_paused
    print("Watcher started.")
    state = load_state()

    # Приветствие в канал при рестарте скрипта
    try:
        bot.send_message(CHANNEL_ID, "🟢 Бот-монитор перезапущен.", parse_mode="HTML")
    except: pass

    while True:
        # Если пауза - стоим и ждем /go. 
        # ВАЖНО: Пока мы тут стоим, спам в личку ТОЖЕ не идет (так как цикл один).
        # Если хочешь, чтобы спам шел даже во время паузы канала - логика будет сложнее.
        # Сейчас пауза останавливает ВСЁ.
        pause_event.wait()

        try:
            print(f"[{datetime.now().strftime('%H:%M')}] Check...")
            content = download_xlsx(URL)
            
            if content:
                curr_text = extract_text(content)
                curr_hash = hashlib.sha256(curr_text.encode()).hexdigest()
                last_hash = state.get("hash")
                
                users = load_users()
                has_changes = False
                
                # 1. АНАЛИЗ ИЗМЕНЕНИЙ
                if last_hash is None:
                    print("Первый прогон.")
                    save_state(curr_hash, curr_text)
                    state = {"hash": curr_hash, "text": curr_text}
                    # Первый раз в канал не пишем, чтобы не шуметь, или можно написать:
                    # bot.send_message(CHANNEL_ID, "База инициализирована.")
                
                elif curr_hash != last_hash:
                    print("!!! CHANGES DETECTED !!!")
                    has_changes = True
                    diff = make_diff(state.get("text"), curr_text)
                    
                    # === ЛОГИКА ДЛЯ КАНАЛА (Только при изменениях) ===
                    msg_channel = f"🚨 <b>РАСПИСАНИЕ ИЗМЕНИЛОСЬ!</b> (УБ-21)\n\n<pre>{diff}</pre>\n\n⏸ <i>Пауза до команды /go</i>"
                    
                    try:
                        f_chan = BytesIO(content)
                        f_chan.name = f"schedule_NEW.xlsx"
                        m = bot.send_document(CHANNEL_ID, f_chan, caption=msg_channel, parse_mode="HTML")
                        bot.pin_chat_message(CHANNEL_ID, m.message_id)
                    except Exception as e:
                        print(f"Channel Error: {e}")

                    # Сохраняем новое состояние
                    save_state(curr_hash, curr_text)
                    state = {"hash": curr_hash, "text": curr_text}

                # 2. СПАМ В ЛИЧКУ (Всегда, каждый цикл)
                msg_private = "✅ Все стабильно. Изменений нет."
                if has_changes:
                    msg_private = "🚨 <b>ВНИМАНИЕ! ЕСТЬ ИЗМЕНЕНИЯ!</b> (См. канал)"

                for uid in users:
                    try:
                        f_priv = BytesIO(content)
                        f_priv.name = f"schedule_{datetime.now().strftime('%H-%M')}.xlsx"
                        bot.send_document(uid, f_priv, caption=msg_private, parse_mode="HTML")
                    except Exception as e:
                        print(f"User {uid} Error: {e}")

                # 3. АКТИВАЦИЯ ПАУЗЫ (Если были изменения)
                if has_changes:
                    is_paused = True
                    pause_event.clear()
                    print("PAUSED.")

        except Exception as e:
            print(f"Loop Error: {e}")

        if not is_paused:
            time.sleep(POLL_SECONDS)

if __name__ == "__main__":
    t = threading.Thread(target=watcher_loop, daemon=True)
    t.start()
    bot.infinity_polling()
